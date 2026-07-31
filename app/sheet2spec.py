#!/usr/bin/env python3
"""채널시트(.numbers/.xlsx) → 스펙 JSON 자동 변환
2026-07-21 확정 컨벤션을 규칙으로 코드화. 표준 시트는 무인 변환, 특수 케이스는 Claude 대화로.
사용: python3 sheet2spec.py 시트경로 [스펙출력.json] [--name YYMMDD_행사명]
"""
import sys, os, json, re, datetime, unicodedata

STEREO_WIDTH = {'ag 1': 65.75, 'ag1': 65.75, 'ag 2': 64.5, 'ag2': 64.5,
                'piano': 61.0, 'key 1': 64.75, 'key1': 64.75, 'key 2': 58.0, 'key2': 58.0,
                'eg': 55.75, 'eg 1': 55.75, 'eg 2': 55.75}
DEFAULT_WIDTH = 60.0
TOM_POS = [41, 0, -40, -95]

ROMAN = {'가':'ga','간':'gan','갈':'gal','감':'gam','강':'gang','경':'gyeong','고':'go','구':'gu','국':'guk',
 '권':'kwon','규':'gyu','근':'geun','금':'geum','기':'gi','김':'kim','나':'na','남':'nam','노':'no','다':'da',
 '대':'dae','도':'do','동':'dong','라':'ra','람':'ram','래':'rae','령':'ryeong','루':'lu','리':'ri','린':'rin',
 '마':'ma','명':'myeong','무':'mu','문':'mun','민':'min','박':'park','배':'bae','백':'baek','범':'beom',
 '별':'byeol','보':'bo','봉':'bong','빈':'bin','상':'sang','서':'seo','석':'seok','선':'seon','성':'seong',
 '세':'se','소':'so','솔':'sol','수':'su','숙':'suk','순':'sun','슬':'seul','승':'seung','시':'si','신':'shin',
 '아':'a','안':'an','애':'ae','양':'yang','어':'eo','언':'eon','에':'e','여':'yeo','연':'yeon','영':'young',
 '예':'ye','오':'o','온':'on','용':'yong','우':'woo','운':'un','원':'won','유':'yu','윤':'yun','율':'yul',
 '은':'eun','음':'eum','이':'lee','인':'in','일':'il','임':'im','자':'ja','장':'jang','재':'jae','전':'jeon',
 '정':'jung','제':'je','조':'jo','종':'jong','주':'ju','준':'jun','지':'ji','진':'jin','찬':'chan','창':'chang',
 '채':'chae','천':'cheon','철':'cheol','최':'choi','태':'tae','하':'ha','한':'han','해':'hae','헌':'heon',
 '혁':'hyeok','현':'hyun','형':'hyung','혜':'hye','호':'ho','홍':'hong','효':'hyo','후':'hu','훈':'hun',
 '희':'hee','그':'geu','싱':'sing','늘':'neul','랑':'rang','솜':'som','든':'deun','행':'haeng','람':'ram','송':'song','시':'si',
 '헤':'he','늘':'neul','다':'da','은':'eun','군':'gun','산':'san','익':'ik','광':'gwang','부':'bu','울':'ul','청':'cheong','천':'cheon','팀':'tim','룩':'look','프':'peu','티':'ti','드':'deu','스':'seu','더':'deo','집':'jip','회':'hoe','련':'ryeon','수':'su','제':'je','열':'yeol','방':'bang','학':'hak','관':'gwan','사':'sa','공':'gong','장':'jang','트':'teu','홀':'hol','록':'rok'}


WORDS = {'구즈넥': 'Goose', '싱어': 'Sing', '스피치': 'Speech', '설교자': 'Pastor', '설교': 'Pastor', '사회자': 'MC', '인도자': 'Leader',
         '연출': 'Dir', '예비': 'Spare', '드럼': 'Drum', '합창': 'Choir', '코러스': 'Chorus',
         '소리': 'Sori', '클릭': 'Click', '마스터건': 'MasterGun', '에스더': 'Esther', '기프티드': 'Gftid', '제이어스': 'JUS', '뉴스공장': 'NewsFactory', '수련회': 'Retreat', '고희안': 'KHA', '홀리그라운드': 'HolyGround', '롤링홀': 'RollingHall', '백석예대': 'BSU'}



# 개정 로마자(RR) 자모 테이블 — ROMAN 사전에 없는 음절의 폴백 (한글 깨짐 방지)
_CHO = ['g','kk','n','d','tt','r','m','b','pp','s','ss','','j','jj','ch','k','t','p','h']
_JUNG = ['a','ae','ya','yae','eo','e','yeo','ye','o','wa','wae','oe','yo','u','wo','we','wi','yu','eu','ui','i']
_JONG = ['','k','k','k','n','n','n','t','l','k','m','p','l','l','p','l','m','p','p','t','t','ng','t','t','k','t','p','t']


def _syllable_rr(ch):
    code = ord(ch) - 0xAC00
    cho, rest = divmod(code, 588)
    jung, jong = divmod(rest, 28)
    return _CHO[cho] + _JUNG[jung] + _JONG[jong]

def romanize(s):
    s = unicodedata.normalize('NFC', s)
    for k, v in WORDS.items():
        s = s.replace(k, v)
    out = []
    for ch in s:
        if ord(ch) < 128:
            out.append(ch)
        elif ch in ROMAN:
            out.append(ROMAN[ch])
        elif 0xAC00 <= ord(ch) <= 0xD7A3:
            out.append(_syllable_rr(ch))
        else:
            out.append('')
    r = ''.join(out)
    return (r[:1].upper() + r[1:]) if r else 'Ch'


def is_ascii(s):
    return all(ord(c) < 128 for c in s)


def read_rows(path):
    rows = []
    if path.lower().endswith('.numbers'):
        from numbers_parser import Document
        doc = Document(path)
        t = doc.sheets[0].tables[0]
        for row in t.rows():
            rows.append(['' if c.value is None else str(c.value) for c in row])
    else:
        from openpyxl import load_workbook
        ws = load_workbook(path).active
        for row in ws.iter_rows():
            rows.append(['' if c.value is None else str(c.value) for c in row])
    return rows


def parse(path):
    rows = read_rows(path)
    hdr = None
    for i, r in enumerate(rows):
        low = [str(x).strip().lower() for x in r]
        if 'input' in low and 'name' in low:
            hdr = i
            cols = {name: low.index(name) for name in ('input', 'name') if name in low}
            cols['io'] = low.index('io input') if 'io input' in low else cols['name'] - 1
            out_cols = {}
            if 'output' in low:
                oi = low.index('output')
                out_cols = {'device': oi - 1, 'output': oi, 'name': oi + 1}
            break
    assert hdr is not None, '헤더 행을 찾지 못함'

    chans = []          # (ch, name, io)
    for r in rows[hdr + 1:]:
        try:
            ch = int(float(r[0]))
        except (ValueError, IndexError):
            continue
        name = str(r[cols['name']]).strip().replace('\n', ' ')
        io = str(r[cols['io']]).strip()
        chans.append((ch, name, io))

    outputs = []        # (device, output, name)
    if out_cols:
        for r in rows[hdr + 1:]:
            try:
                dev = str(r[out_cols['device']]).strip()
                out = str(r[out_cols['output']]).strip()
                nm = str(r[out_cols['name']]).strip().replace('\n', ' ')
            except IndexError:
                continue
            if out:
                outputs.append((dev, out, nm))
    return chans, outputs


def classify(chans):
    """이름 규칙으로 그룹·DCA·팬 결정 + 스테레오 페어 탐지"""
    named = {ch: nm for ch, nm, io in chans if nm}
    ios = {ch: io for ch, nm, io in chans}
    pairs = []
    used = set()
    seq = sorted(named)
    # 페어: 이름 있는 행 + 다음 행이 IO는 있는데 이름이 빈 경우 (스테레오 후보 악기만)
    STEREO_OK = re.compile(r'piano|key|synth|pad|mtr|eg|gtr|guitar|ag|oh|spd|ppr|프리젠터|drum(?!\s*click)|bgm|fx|sov|mac|music|스테레오|st', re.I)
    all_ch = {ch for ch, nm, io in chans}
    for ch in seq:
        nxt = ch + 1
        if (nxt in all_ch and nxt not in named and ios.get(nxt, '') and ch not in used
                and STEREO_OK.search(named[ch]) and not re.search(r'click|클릭|ambi|\btb\b|^wl\s*\d+$', named[ch].strip(), re.I)):
            pairs.append((ch, nxt))
            used.update((ch, nxt))
    # L/R 표기 인접 페어 (양쪽 모두 이름 있는 행)
    LR_L = re.compile(r'^(.*?)[\s\-_]+[Ll](?:eft)?$')
    LR_R = re.compile(r'^(.*?)[\s\-_]+[Rr](?:ight)?$')
    for ch in seq:
        nxt = ch + 1
        if ch in used or nxt in used or nxt not in named:
            continue
        m1, m2 = LR_L.match(named[ch].strip()), LR_R.match(named[nxt].strip())
        if m1 and m2 and m1.group(1) and m1.group(1).lower() == m2.group(1).lower():
            pairs.append((ch, nxt))
            used.update((ch, nxt))
    # OH Ride / OH HH 인접 페어 관례
    for ch in seq:
        nm = named[ch].lower()
        if 'oh' in nm and 'ride' in nm and (ch + 1) in named and 'hh' in named[ch + 1].lower() and ch not in used:
            pairs.append((ch, ch + 1))
            used.update((ch, ch + 1))

    pairmap = {}
    for a, b in pairs:
        pairmap[a] = b
        pairmap[b] = a
    lr_first = {a for a, b in pairs
                if a in named and named.get(b, '') and LR_L.match(named[a].strip())}

    channels = []
    tom_chs = [c for c in seq if re.match(r'tom\s*\d', named[c].lower())]
    tom_positions = TOM_POS if len(tom_chs) != 2 else [41, -40]   # 재즈킷: Tom2 → Tom3 자리

    for ch in seq:
        nm = named[ch]
        partner = pairmap.get(ch)
        base = named.get(min(ch, partner) if partner else ch, nm)
        low = base.lower()
        is_pair = partner is not None
        c = {'ch': ch, 'name': (nm if nm else base)[:12]}

        def st(width):
            return {'width_deg': width}

        if re.search(r'kick|snare', low):
            c.update(group='Drums', dca=['OnAir', 'inst', 'Drums'], pan='mono')
        elif re.fullmatch(r'hh|hi-?hat', low):
            c.update(group='Drums', dca=['OnAir', 'inst', 'Drums'], pan={'pos_deg': 79.5})
        elif re.match(r'tom\s*\d', low):
            idx = tom_chs.index(ch) if ch in tom_chs else 0
            pos = tom_positions[idx] if idx < len(tom_positions) else -95
            c.update(group='Drums', dca=['OnAir', 'inst', 'Drums'], pan={'pos_deg': pos})
        elif 'oh' in low.split() or low.startswith('oh'):
            c.update(group='Drums', dca=['OnAir', 'inst', 'Drums'], pan='oh' if is_pair else 'mono')
            if is_pair and low.startswith('oh'):
                c['klang_name'] = 'OH'
        elif ('spd' in low or 'perc' in low or 'conga' in low or 'congga' in low
              or 'cajon' in low or re.match(r'hand\s*\d', low)):
            c.update(group='Drums', dca=['OnAir', 'inst', 'Drums'],
                     pan=st(DEFAULT_WIDTH) if is_pair else 'mono')
        elif 'bass' in low:
            c.update(group='Bass', dca=['OnAir', 'inst'], pan='mono')
        elif 'click' in low or '클릭' in base:
            c.update(group=None, dca=[], pan='mono')
        elif 'tb' in low.split() or low.startswith('tb') or low.endswith('tb'):
            c.update(group=None, dca=[], pan='mono')
        elif 'ambi' in low:
            c.update(group='AMBI', dca=['AMBI'], pan='mono')
        elif 'piano' in low:
            c.update(group='Piano', dca=['OnAir', 'inst'],
                     pan=st(STEREO_WIDTH.get('piano', DEFAULT_WIDTH)) if is_pair else 'mono')
        elif re.search(r'key|synth|pad|mtr', low):
            w = next((v for k, v in STEREO_WIDTH.items() if k in low), DEFAULT_WIDTH)
            c.update(group='Key', dca=['OnAir', 'inst'],
                     pan=st(w) if is_pair else 'mono')
        elif re.search(r'eg|gtr|guitar', low):
            c.update(group='GTR', dca=['OnAir', 'inst'],
                     pan=st(STEREO_WIDTH.get('eg', DEFAULT_WIDTH)) if is_pair else 'mono')
        elif re.search(r'\bag\b|ag\s*\d|어쿠스틱', low):
            w = next((v for k, v in STEREO_WIDTH.items() if k in low), 65.0)
            c.update(group='AG', dca=['OnAir', 'inst'], pan=st(w) if is_pair else 'mono')
        elif re.search(r'ppr|프리젠터|playback|재생', low):
            c.update(group=None, dca=[], pan='mono')
        elif re.search(r'mc|pastor|설교|사회|speech|스피치|예비', low):
            c.update(group='Sings', dca=[], pan='mono')
        elif re.search(r'wl|무선|sing|vox|chorus|합창|코러스|소프라노|leader|인도', low) or not is_ascii(base):
            c.update(group='Sings', dca=['OnAir', 'Sings'], pan='mono')
        else:
            c.update(group=None, dca=[], pan='mono')

        if ch in lr_first and 'klang_name' not in c:
            m_lr = LR_L.match(c['name'].strip())
            if m_lr and m_lr.group(1):
                c['klang_name'] = m_lr.group(1)[:12]
        if not is_ascii(c.get('klang_name', c['name'])):
            m = re.match(r'^(\d+)\s*(.+)$', c['name'])
            body = m.group(2) if m else c['name']
            prefix = (m.group(1) + ' ') if m else ''
            c['klang_name'] = (prefix + romanize(body))[:12]
        channels.append(c)
    return channels, pairs


def parse_outputs(outputs):
    mixes, mix_pairs, matrix = {}, [], {}
    iem_mixes = []
    for dev, out, nm in outputs:
        m = re.match(r'mix\s*(\d+)', out.lower())
        if m:
            n = int(m.group(1))
            if 'iem' in dev.lower() or 'iem' in nm.lower():
                iem_mixes.append((n, nm or f'IEM {len(iem_mixes)+1}'))
            elif nm:
                label = nm if is_ascii(nm) else nm
                mixes[str(n)] = {'name': label[:12]}
        m2 = re.match(r'mtrx\s*(\d+)', out.lower())
        if m2 and nm:
            n = int(m2.group(1))
            base = re.sub(r'\s*[LR]$', '', nm)
            link = 'L' if nm.strip().endswith('L') else ('R' if nm.strip().endswith('R') else None)
            ent = {'name': base[:12]}
            if link:
                ent['link'] = link
            else:
                ent['mono'] = True
            matrix[str(n)] = ent
    for k, (n, nm) in enumerate(iem_mixes):
        label = nm if is_ascii(nm) else f'IEM {k+1}'
        mixes[str(n)] = {'name': label[:12]}
        mixes[str(n + 1)] = {'name': label[:12]}
        mix_pairs.append([n, n + 1])
    return mixes, mix_pairs, matrix


def build_spec(sheet_path, show_name=None):
    chans, outputs = parse(sheet_path)
    channels, pairs = classify(chans)
    mixes, mix_pairs, matrix = parse_outputs(outputs)
    today = datetime.date.today().strftime('%y%m%d')
    ascii_name = None
    if not show_name:
        base = os.path.splitext(os.path.basename(sheet_path))[0]
        base = unicodedata.normalize('NFC', base)
        base = re.sub(r'^\d{6}[_ ]?', '', base)
        ascii_base = base if is_ascii(base) else romanize(base)
        ascii_name = f'{today}_{re.sub(r"[^A-Za-z0-9 _-]", "", ascii_base).strip().replace(" ", "") or "Show"}'
        show_name = ascii_name
    groups_used = [g for g in ['Drums', 'Key', 'GTR', 'AG', 'Sings', 'Piano', 'Bass', 'AMBI']
                   if any(c.get('group') == g for c in channels)]
    ascii_name = ascii_name or (show_name if is_ascii(show_name) else romanize(show_name))
    spec = {
        'name': show_name,
        'ascii_name': ascii_name,
        'snapshot': ascii_name.replace('_', '')[:10],
        'groups': groups_used,
        'channels': channels,
        'pairs': [list(p) for p in pairs],
        'mixes': mixes,
        'mix_pairs': mix_pairs,
        'matrix': matrix,
    }
    if any('AMBI' in c.get('dca', []) for c in channels):
        spec['dca_names'] = {'12': 'AMBI'}
    return spec


if __name__ == '__main__':
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    name = None
    for i, a in enumerate(sys.argv):
        if a == '--name':
            name = sys.argv[i + 1]
    spec = build_spec(args[0], name)
    out = args[1] if len(args) > 1 else '/tmp/showfile_spec.json'
    json.dump(spec, open(out, 'w'), ensure_ascii=False, indent=1)
    print(out)
    print(f"채널 {len(spec['channels'])}개, 페어 {len(spec['pairs'])}개, 믹스 {len(spec['mixes'])}개, 그룹 {spec['groups']}")
